#!/usr/bin/env python3
"""Actualiza perfiles y suscripciones desde listado_carnets.xlsx."""

from __future__ import annotations

import json
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
MCP_CONFIG = ROOT / ".cursor" / "mcp.json"
PROJECT_REF = "evtgfwphownszsjrkpph"
API_URL = f"https://api.supabase.com/v1/projects/{PROJECT_REF}/database/query"
DEFAULT_XLSX = ROOT.parent / "DB" / "listado_carnets.xlsx"
OUT_DIR = ROOT / "supabase" / "seed" / "generated_import" / "carnets_update"
TODAY = date(2026, 6, 10)

SOCIO_FROM_TEXT_RE = re.compile(r"LIBRE\d*-0*(\d+)(?:\s|$)", re.I)
SOCIO_DASH_RE = re.compile(r"-0*(\d+)\s{2,}[A-ZÁÉÍÓÚÑ]", re.I)
SOCIO_END_RE = re.compile(r"-0*(\d+)\s*$")
ESTADO_PRIORITY = {"Vigente": 3, "Suspendido": 2, "Vencido": 1, "Anulado": 0}

PLAN_ALIASES = {
    "DESC. DE ESTUDIANTES 30D.LIBRE-": "DESC. DE ESTUDIANTES 30D.LIBRE-",
    "CROSSFIT 30D.LIBRE-": "CROSSFIT 30D.LIBRE-",
    "ESCUELA DE PADEL MP 30D.LIBRE-": "ESCUELA DE PADEL MP 30D.LIBRE-",
    "PROMO MES MAYO 30D.LIBRE-": "PROMO MES MAYO 30D.LIBRE-",
    "MUSCULACION 6D.LIBRE-": "MUSCULACION 6D.LIBRE-",
    "MUSCULACION 90D.LIBRE- 3 M": "MUSCULACION 90D.LIBRE- 3 M",
    "MUSCULACION 30D.LIBRE-": "MUSCULACION 30D.LIBRE-",
}


def load_token() -> str:
    token = __import__("os").environ.get("SUPABASE_ACCESS_TOKEN")
    if token:
        return token
    data = json.loads(MCP_CONFIG.read_text())
    args = data["mcpServers"]["supabase-dual-gym"]["args"]
    return args[args.index("--access-token") + 1]


def run_sql(token: str, sql: str) -> None:
    body = json.dumps({"query": sql}).encode()
    req = urllib.request.Request(
        API_URL,
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "dualgym-carnets-update/1.0",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=300) as resp:
        if resp.status not in (200, 201):
            raise RuntimeError(f"Unexpected status {resp.status}: {resp.read()}")


def to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def normalize_name(name: str | None) -> str:
    if not name:
        return ""
    return re.sub(r"\s+", " ", name.strip().upper())


def extract_socio(*texts: str | None) -> str | None:
    for text in texts:
        if not text:
            continue
        for rx in (SOCIO_FROM_TEXT_RE, SOCIO_DASH_RE, SOCIO_END_RE):
            match = rx.search(str(text))
            if match:
                return str(int(match.group(1)))
    return None


def extract_plan_and_name_from_merged(text: str) -> tuple[str | None, str | None, str | None]:
    text = re.sub(r"\s+", " ", str(text).strip())
    match = re.match(r"^(.+?\d+D[\d.]*LIBRE)-(.+)$", text, re.I)
    if not match:
        return text, None, None
    left, right = match.group(1).strip(), match.group(2).strip()
    socio = extract_socio(left, right, text)
    if socio:
        name = re.sub(rf"^{re.escape(socio)}\s*", "", right).strip()
        if not name:
            name = re.sub(rf".*-0*{socio}\s*", "", text).strip()
        plan = re.sub(rf"-?0*{socio}\s*$", "", left).strip()
        return plan, name or None, socio
    return left, right, None


def normalize_plan(raw: str | None) -> str:
    if not raw:
        return "MUSCULACION 30D.LIBRE-"
    text = re.sub(r"\s+", " ", raw.upper().strip())
    text = text.replace("LI1B0R", "LIBRE").replace("LI1B0RE", "LIBRE")
    text = re.sub(r"LIBRE\d*-0*\d+.*$", "LIBRE-", text)
    text = re.sub(r"LIBRE-0*\d+\s*$", "LIBRE-", text)
    if "ESTUDIANTES" in text:
        return "DESC. DE ESTUDIANTES 30D.LIBRE-"
    if "CROSSFIT" in text:
        return "CROSSFIT 30D.LIBRE-"
    if "PADEL" in text:
        return "ESCUELA DE PADEL MP 30D.LIBRE-"
    if "PROMO MES MAYO" in text:
        return "PROMO MES MAYO 30D.LIBRE-"
    if re.search(r"\b6D\b", text):
        return "MUSCULACION 6D.LIBRE-"
    if re.search(r"\b90D\b", text):
        return "MUSCULACION 90D.LIBRE- 3 M"
    return "MUSCULACION 30D.LIBRE-"


def membership_from_estado(estado: str, vence: date | None) -> tuple[str, str]:
    if estado == "Vigente" and vence and vence >= TODAY:
        return "active", "active"
    if estado == "Suspendido":
        return "inactive", "unpaid"
    return "inactive", "canceled"


def parse_row(row) -> dict | None:
    if not row or not any(cell is not None and str(cell).strip() for cell in row):
        return None
    first = str(row[0] or "")
    if "Fecha" in first and "Carnet" in first:
        return None
    if "Listado de Carnets" in first:
        return None
    if not isinstance(row[0], (datetime, date)):
        return None

    estado = str(row[8]).strip() if len(row) > 8 and row[8] else ""
    if not estado:
        return None

    c2 = str(row[2] or "").strip() if len(row) > 2 else ""
    c3 = str(row[3] or "").strip() if len(row) > 3 else ""
    c4 = str(row[4] or "").strip() if len(row) > 4 else ""
    plan = nombre = numero_socio = None

    if c4:
        plan = f"{c2} {c3}".strip() if c3 else c2
        nombre = c4
        numero_socio = extract_socio(c3, c2, plan)
    elif c2:
        plan, nombre, numero_socio = extract_plan_and_name_from_merged(c2)

    return {
        "fecha": to_date(row[0]),
        "nro_carnet": int(row[1]) if isinstance(row[1], (int, float)) else None,
        "numero_socio": numero_socio,
        "plan_raw": plan,
        "plan_name": normalize_plan(plan),
        "nombre": nombre,
        "nombre_norm": normalize_name(nombre),
        "vence": to_date(row[5]) if len(row) > 5 else None,
        "estado": estado,
    }


def pick_carnet(rows: list[dict]) -> dict:
    def sort_key(row: dict) -> tuple:
        vence = row["vence"] or date.min
        fecha = row["fecha"] or date.min
        return (ESTADO_PRIORITY.get(row["estado"], 0), vence, fecha)

    return max(rows, key=sort_key)


def parse_workbook(path: Path) -> tuple[dict[str, dict], list[dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parsed: list[dict] = []
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            record = parse_row(row)
            if record:
                parsed.append(record)

    by_socio: dict[str, list[dict]] = defaultdict(list)
    no_socio: list[dict] = []
    for record in parsed:
        if record["numero_socio"]:
            by_socio[record["numero_socio"]].append(record)
        else:
            no_socio.append(record)

    selected = {socio: pick_carnet(rows) for socio, rows in by_socio.items()}
    return selected, no_socio


def sql_text(value: str | None) -> str:
    if value is None:
        return "NULL"
    cleaned = value.strip()
    if not cleaned:
        return "NULL"
    return "'" + cleaned.replace("'", "''") + "'"


def sql_date(value: date | None) -> str:
    return "NULL" if value is None else f"'{value.isoformat()}'"


def build_name_index(profiles: list[dict]) -> dict[str, str]:
    by_name: dict[str, str] = {}
    for profile in profiles:
        keys = {
            normalize_name(f"{profile['last_name']}, {profile['first_name']}"),
            normalize_name(f"{profile['last_name']} {profile['first_name']}"),
            normalize_name(f"{profile['first_name']} {profile['last_name']}"),
        }
        for key in keys:
            if key:
                by_name[key] = profile["numero_socio"]
    return by_name


def build_updates(selected: dict[str, dict], no_socio: list[dict], profiles: list[dict]) -> list[dict]:
    by_name = build_name_index(profiles)
    records_by_profile: dict[str, list[dict]] = defaultdict(list)

    for record in list(selected.values()) + no_socio:
        profile_socio = by_name.get(record["nombre_norm"])
        target_socio = profile_socio or record.get("numero_socio")
        if not target_socio:
            continue
        enriched = dict(record)
        enriched["target_socio"] = target_socio
        enriched["match"] = "nombre" if profile_socio else "numero_socio"
        records_by_profile[target_socio].append(enriched)

    updates: dict[str, dict] = {}
    for socio, rows in records_by_profile.items():
        record = pick_carnet(rows)
        membership_status, sub_status = membership_from_estado(record["estado"], record["vence"])
        updates[socio] = {
            "numero_socio": socio,
            "numero_carnet": str(record["nro_carnet"]) if record["nro_carnet"] is not None else None,
            "membership_status": membership_status,
            "plan_name": record["plan_name"],
            "fecha_inicio": record["fecha"],
            "vence": record["vence"],
            "sub_status": sub_status,
            "estado": record["estado"],
            "match": record["match"],
        }

    return sorted(updates.values(), key=lambda item: int(item["numero_socio"]))


def build_batch_sql(updates: list[dict]) -> str:
    if not updates:
        return ""

    values = []
    for item in updates:
        values.append(
            "("
            f"{sql_text(item['numero_socio'])}, "
            f"{sql_text(item['numero_carnet'])}, "
            f"{sql_text(item['membership_status'])}, "
            f"{sql_text(item['plan_name'])}, "
            f"{sql_date(item['fecha_inicio'])}, "
            f"{sql_date(item['vence'])}, "
            f"{sql_text(item['sub_status'])}, "
            f"{sql_text(item['estado'])}"
            ")"
        )

    return f"""
CREATE TEMP TABLE carnet_updates (
  numero_socio text PRIMARY KEY,
  numero_carnet text,
  membership_status text,
  plan_name text,
  fecha_inicio date,
  vence date,
  sub_status text,
  estado text
) ON COMMIT DROP;

INSERT INTO carnet_updates (numero_socio, numero_carnet, membership_status, plan_name, fecha_inicio, vence, sub_status, estado)
VALUES
  {",\n  ".join(values)};

UPDATE public.profiles p
SET
  numero_carnet = c.numero_carnet,
  membership_status = c.membership_status,
  updated_at = now()
FROM carnet_updates c
WHERE p.numero_socio = c.numero_socio;

UPDATE public.subscriptions s
SET
  status = c.sub_status,
  current_period_start = COALESCE(c.fecha_inicio, s.current_period_start),
  current_period_end = COALESCE(c.vence, s.current_period_end),
  plan_id = COALESCE(
    (SELECT id FROM public.plans WHERE name = c.plan_name LIMIT 1),
    s.plan_id
  ),
  updated_at = now()
FROM carnet_updates c
JOIN public.profiles p ON p.numero_socio = c.numero_socio
WHERE s.profile_id = p.id
  AND s.id = (
    SELECT s2.id
    FROM public.subscriptions s2
    WHERE s2.profile_id = p.id
    ORDER BY s2.created_at DESC
    LIMIT 1
  );
"""


def fetch_profiles(token: str) -> list[dict]:
    sql = (
        "SELECT numero_socio, first_name, last_name "
        "FROM public.profiles "
        "WHERE numero_socio IS NOT NULL "
        "AND created_by_email = 'import@dualgym.local';"
    )
    body = json.dumps({"query": sql}).encode()
    req = urllib.request.Request(
        API_URL,
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "dualgym-carnets-update/1.0",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        payload = json.loads(resp.read().decode())
    return payload


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"No existe {xlsx_path}")
        return 1

    token = load_token()
    print(">> Aplicando migración numero_carnet...")
    migration = (ROOT / "supabase/migrations/20260610120000_add_numero_carnet_to_profiles.sql").read_text()
    run_sql(token, migration)

    print(f">> Parseando {xlsx_path.name}...")
    selected, no_socio = parse_workbook(xlsx_path)
    profiles = fetch_profiles(token)
    updates = build_updates(selected, no_socio, profiles)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = {
        "parsed_socios": len(selected),
        "no_socio_rows": len(no_socio),
        "updates": len(updates),
        "active": sum(1 for item in updates if item["membership_status"] == "active"),
        "inactive": sum(1 for item in updates if item["membership_status"] == "inactive"),
        "matched_by_name": sum(1 for item in updates if item["match"] == "nombre"),
    }
    (OUT_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2))
    (OUT_DIR / "updates.json").write_text(json.dumps(updates, indent=2, ensure_ascii=False, default=str))

    batch_size = 120
    batches = [updates[i : i + batch_size] for i in range(0, len(updates), batch_size)]
    print(f">> Aplicando {len(batches)} lotes ({len(updates)} socios)...")
    for index, batch in enumerate(batches, start=1):
        sql = build_batch_sql(batch)
        out_file = OUT_DIR / f"batch_{index:03d}.sql"
        out_file.write_text(sql)
        print(f"   lote {index}/{len(batches)} ({len(batch)} socios)", flush=True)
        try:
            run_sql(token, sql)
        except urllib.error.HTTPError as exc:
            print(f"   ERROR HTTP {exc.code}: {exc.read().decode()[:500]}")
            return 1
        except Exception as exc:  # noqa: BLE001
            print(f"   ERROR {exc}")
            return 1
        time.sleep(0.2)

    verify_sql = """
SELECT
  (SELECT count(*) FROM public.profiles WHERE numero_carnet IS NOT NULL) AS con_carnet,
  (SELECT count(*) FROM public.profiles WHERE membership_status = 'active' AND numero_carnet IS NOT NULL) AS activos_carnet,
  (SELECT count(*) FROM public.profiles WHERE membership_status = 'inactive' AND numero_carnet IS NOT NULL) AS inactivos_carnet;
"""
    body = json.dumps({"query": verify_sql}).encode()
    req = urllib.request.Request(
        API_URL,
        data=body,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "dualgym-carnets-update/1.0",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        result = json.loads(resp.read().decode())

    print(">> Verificación:", json.dumps(result, indent=2))
    print(">> Manifest:", json.dumps(manifest, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
